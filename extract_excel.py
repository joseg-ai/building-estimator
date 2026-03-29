"""
Extract all data from the building estimator .xlsm workbook.
Dumps each sheet's contents to a text file for analysis.
"""
import openpyxl
import os
import json

WORKBOOK_PATH = r"C:\Users\joseguajardo\Code\building-estimator\26-0325855  GV - SR PEMB 23130 Tomball Pkwy Bldg 14.xlsm"
OUTPUT_DIR = r"C:\Users\joseguajardo\Code\building-estimator\extracted_data"

os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True, keep_vba=False)

print(f"Workbook loaded: {len(wb.sheetnames)} sheets found")
print(f"Sheet names: {wb.sheetnames}")
print()

summary = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows_data = []
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0
    non_empty_cells = 0

    for row in ws.iter_rows(min_row=1, max_row=min(max_row, 200), values_only=False):
        row_vals = []
        for cell in row:
            val = cell.value
            if val is not None:
                non_empty_cells += 1
            row_vals.append(val)
        # Only keep rows that have at least one non-None value
        if any(v is not None for v in row_vals):
            rows_data.append((row[0].row, row_vals))

    summary[sheet_name] = {
        "max_row": max_row,
        "max_col": max_col,
        "non_empty_cells": non_empty_cells,
        "data_rows": len(rows_data),
    }

    # Write to file
    safe_name = sheet_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    out_path = os.path.join(OUTPUT_DIR, f"{safe_name}.txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"Sheet: {sheet_name}\n")
        f.write(f"Dimensions: {max_row} rows x {max_col} cols\n")
        f.write(f"Non-empty cells: {non_empty_cells}\n")
        f.write(f"Data rows (non-blank): {len(rows_data)}\n")
        f.write("=" * 80 + "\n\n")

        for row_num, vals in rows_data:
            # Trim trailing Nones
            while vals and vals[-1] is None:
                vals = vals[:-1]
            formatted = []
            for i, v in enumerate(vals):
                if v is not None:
                    formatted.append(f"[{i+1}]={v}")
            if formatted:
                f.write(f"Row {row_num}: {' | '.join(formatted)}\n")

    print(f"  {sheet_name}: {max_row}r x {max_col}c, {non_empty_cells} cells, {len(rows_data)} data rows -> {safe_name}.txt")

# Write summary
with open(os.path.join(OUTPUT_DIR, "_summary.json"), "w") as f:
    json.dump(summary, f, indent=2)

print(f"\nAll data extracted to: {OUTPUT_DIR}")
